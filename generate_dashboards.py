#!/usr/bin/env python3
"""
generate_dashboards.py

Regenerates the embedded POSITIONS data in both dashboards from the master
Google Sheet export (xlsx). Replaces auto_refresh.py from the old pipeline
(that script was not available for this handover, so this one was rebuilt
to do the same job against the current sheet layout).

Usage:
    python3 generate_dashboards.py --xlsx "/path/to/B2C x Alma Available & Open Positions.xlsx"

What it does:
    1. Opens the workbook (twice: once with formulas visible, once with
       cached values, so the Brochure hyperlink can be recovered either
       way it was entered in Google Sheets).
    2. For each of the two sheets below, finds the header row automatically,
       reads every data row after it, and maps columns onto that sheet's
       field schema (the two sheets now use different column sets).
    3. Injects the resulting data as a JSON array into the matching
       dashboard's `const POSITIONS = [...]` block.
    4. Writes the two index.html files in place, ready to deploy.

Sheet -> dashboard mapping:
    "Alma Positions | Ausbildung"       -> ausbildung/index.html
      (English/German headers: Year, Position ID, Official Title, Y1-Y4
      Salary, Program duration, etc.)
    "Alma Positions | 18b, 19c, 16d"    -> 18b-19c-16d/index.html
      (Vietnamese headers: Năm, ID, Vị trí, Brochure, Lương cứng
      trước/sau thuế, etc. — no year-by-year salary, no program duration,
      these are direct-employment positions, not apprenticeships.)

Rows with no ID/Position ID AND no title are skipped (stray rows).
"""

import argparse
import json
import re
import sys
from pathlib import Path

import openpyxl

SCRIPT_DIR = Path(__file__).resolve().parent

AUSBILDUNG_SHEET = "Alma Positions | Ausbildung"
AUSBILDUNG_HTML = SCRIPT_DIR / "ausbildung" / "index.html"

FACHKRAFT_SHEET = "Alma Positions | 18b, 19c, 16d"
FACHKRAFT_HTML = SCRIPT_DIR / "18b-19c-16d" / "index.html"


def norm(text):
    if text is None:
        return ""
    return str(text).replace("\n", " ").replace("\r", " ").strip().lower()


def clean_num(value):
    if value is None:
        return ""
    s = str(value).strip()
    if s.endswith(".0"):
        return s[:-2]
    return s


def money(value):
    if value is None:
        return ""
    s = str(value).strip()
    if not s or set(s) == {"-"}:
        return ""
    try:
        f = float(s)
        return "€" + format(int(round(f)), ",")
    except ValueError:
        return s


def category_of(industry):
    i = (industry or "").lower()
    if any(k in i for k in [
        "gesundheit", "pflege", "medizin", "zahn", "klinik",
        "y tế", "điều dưỡng", "chăm sóc", "y tá", "care",
    ]):
        return "healthcare"
    if any(k in i for k in [
        "handel", "service", "verkauf", "gastro", "hotel",
        "nhà hàng", "dịch vụ", "ẩm thực", "bán lẻ", "khách sạn",
        "nấu ăn", "đầu bếp", "bếp", "retail", "restaurant",
    ]):
        return "services"
    if any(k in i for k in [
        "technik", "druck", "handwerk", "industrie", "metall", "elektro",
        "kỹ thuật", "logistik", "kho vận", "hậu cần", "sản xuất", "cơ khí",
        "logistics", "warehouse",
    ]):
        return "trade"
    # Unmatched industries default to "services" (the most general bucket)
    # rather than silently landing under healthcare.
    return "services"


def find_header_row(ws, required_tokens, max_scan=15):
    """Find the row that contains all of `required_tokens` (normalized substrings)."""
    for r in range(1, min(ws.max_row, max_scan) + 1):
        row_values = [norm(ws.cell(row=r, column=c).value) for c in range(1, ws.max_column + 1)]
        joined = " | ".join(row_values)
        if all(any(tok in v for v in row_values) or tok in joined for tok in required_tokens):
            return r
    raise ValueError(f"Could not find a header row containing: {required_tokens}")


def map_headers(ws, header_row, field_matchers):
    headers = [ws.cell(row=header_row, column=c).value for c in range(1, ws.max_column + 1)]
    normed = [norm(h) for h in headers]
    field_to_col = {}
    for field, test in field_matchers:
        for i, h in enumerate(normed):
            if h and field not in field_to_col and test(h):
                field_to_col[field] = i + 1  # 1-indexed column
                break
    return field_to_col


# ---------------------------------------------------------------------------
# Ausbildung sheet (unchanged schema)
# ---------------------------------------------------------------------------

AUSBILDUNG_FIELD_MATCHERS = [
    ("year", lambda h: h == "year"),
    ("position_id", lambda h: "position id" in h),
    ("status", lambda h: h == "status"),
    ("industry", lambda h: h == "industry"),
    ("title", lambda h: "official title" in h),
    ("company_type", lambda h: "company type" in h),
    ("location", lambda h: h == "location"),
    ("state", lambda h: h == "state"),
    ("requirements", lambda h: "key student requirements" in h),
    ("notes", lambda h: "additional notes" in h),
    ("language_level", lambda h: "language lvl requirement" in h),
    ("season", lambda h: h == "season"),
    ("start_date", lambda h: "start date" in h),
    ("submission_deadline", lambda h: "submission deadline" in h),
    ("duration", lambda h: "program duration" in h),
    ("y1_salary", lambda h: "y1" in h and "salary" in h),
    ("y2_salary", lambda h: "y2" in h and "salary" in h),
    ("y3_salary", lambda h: "y3" in h and "salary" in h),
    ("y4_salary", lambda h: "y4" in h and "salary" in h),
    ("fulltime_salary", lambda h: "full-time salary" in h or "full time salary" in h),
    ("housing", lambda h: h == "housing"),
    ("rent_pricing", lambda h: "rent pricing" in h),
    ("vacation_days", lambda h: "vacation days" in h),
    ("night_shift", lambda h: "night shift" in h),
    ("other_benefits", lambda h: "other financial benefits" in h),
    ("deutschlandticket", lambda h: "deutschlandticket" in h),
    ("extra_perks", lambda h: "extra perks" in h),
]


def extract_ausbildung_rows(ws):
    header_row = find_header_row(ws, required_tokens=["year", "position id"])
    field_to_col = map_headers(ws, header_row, AUSBILDUNG_FIELD_MATCHERS)

    def get(row, field):
        col = field_to_col.get(field)
        if col is None:
            return ""
        v = ws.cell(row=row, column=col).value
        return "" if v is None else str(v).strip()

    rows = []
    for r in range(header_row + 1, ws.max_row + 1):
        pid = get(r, "position_id")
        title = get(r, "title")
        if not pid and not title:
            continue
        rows.append({
            "year": clean_num(get(r, "year")),
            "position_id": pid,
            "status": get(r, "status"),
            "title": title,
            "company_type": get(r, "company_type"),
            "location": get(r, "location"),
            "state": get(r, "state"),
            "requirements": get(r, "requirements"),
            "notes": get(r, "notes"),
            "language_level": get(r, "language_level"),
            "season": get(r, "season"),
            "start_date": get(r, "start_date"),
            "submission_deadline": get(r, "submission_deadline"),
            "duration": clean_num(get(r, "duration")),
            "y1_salary": money(get(r, "y1_salary")),
            "y2_salary": money(get(r, "y2_salary")),
            "y3_salary": money(get(r, "y3_salary")),
            "y4_salary": money(get(r, "y4_salary")),
            "fulltime_salary": get(r, "fulltime_salary"),
            "housing": get(r, "housing"),
            "rent_pricing": get(r, "rent_pricing"),
            "vacation_days": clean_num(get(r, "vacation_days")),
            "night_shift": get(r, "night_shift"),
            "other_benefits": get(r, "other_benefits"),
            "deutschlandticket": get(r, "deutschlandticket"),
            "extra_perks": get(r, "extra_perks"),
            "category": category_of(get(r, "industry")),
        })
    return rows


# ---------------------------------------------------------------------------
# §18b, 19c, 16d sheet (new Vietnamese schema — direct employment, no
# year-by-year salary, no program duration/deadline, adds Brochure link)
# ---------------------------------------------------------------------------

FACHKRAFT_FIELD_MATCHERS = [
    ("year", lambda h: h == "năm"),
    ("date_posted", lambda h: "ngày đăng" in h),
    ("position_id", lambda h: h == "id"),
    ("status", lambda h: "trạng" in h),  # "Tinh trạng" / "Tình trạng"
    ("industry", lambda h: h == "ngành"),
    ("title", lambda h: h == "vị trí"),
    ("brochure_url", lambda h: h == "brochure"),
    ("company_type", lambda h: "loại hình doanh nghiệp" in h),
    ("city", lambda h: h == "thành phố"),
    ("state", lambda h: h == "bang"),
    ("employer_origin", lambda h: "chủ doanh nghiệp" in h),
    ("requirements", lambda h: "yêu cầu chính" in h),
    ("notes", lambda h: "additional notes" in h),
    ("language_level", lambda h: "trình độ tiếng đức" in h),
    ("start_date", lambda h: "thời gian bắt đầu" in h),
    ("salary_pretax", lambda h: "lương cứng" in h and "trước thuế" in h),
    ("salary_posttax", lambda h: "lương cứng" in h and "sau thuế" in h),
    ("benefits_notes", lambda h: "phúc lợi" in h),
    ("housing", lambda h: h == "nhà ở"),
    ("rent_price", lambda h: "giá thuê nhà" in h),
    ("working_hours", lambda h: "thời gian làm việc" in h),
    ("day_off", lambda h: h == "ngày nghỉ"),
    ("vacation_days", lambda h: "số ngày nghỉ phép" in h),
    ("cost_of_living", lambda h: "chi phí sinh hoạt" in h),
]

HYPERLINK_FORMULA_RE = re.compile(r'HYPERLINK\(\s*"([^"]+)"', re.IGNORECASE)


def get_brochure_url(ws_values, ws_formulas, row, col):
    """Recover the Brochure URL regardless of how it was entered in Sheets:
    a native inserted link, a =HYPERLINK() formula, or plain URL text."""
    if col is None:
        return ""

    cell_f = ws_formulas.cell(row=row, column=col)
    if cell_f.hyperlink and cell_f.hyperlink.target:
        return cell_f.hyperlink.target.strip()

    if isinstance(cell_f.value, str):
        m = HYPERLINK_FORMULA_RE.search(cell_f.value)
        if m:
            return m.group(1).strip()

    cell_v = ws_values.cell(row=row, column=col)
    if isinstance(cell_v.value, str) and cell_v.value.strip().lower().startswith("http"):
        return cell_v.value.strip()

    return ""


def extract_fachkraft_rows(ws_values, ws_formulas):
    header_row = find_header_row(ws_values, required_tokens=["năm", "brochure"])
    field_to_col = map_headers(ws_values, header_row, FACHKRAFT_FIELD_MATCHERS)

    def get(row, field):
        col = field_to_col.get(field)
        if col is None:
            return ""
        v = ws_values.cell(row=row, column=col).value
        return "" if v is None else str(v).strip()

    rows = []
    for r in range(header_row + 1, ws_values.max_row + 1):
        pid = get(r, "position_id")
        title = get(r, "title")
        if not pid and not title:
            continue
        rows.append({
            "year": clean_num(get(r, "year")),
            "date_posted": get(r, "date_posted"),
            "position_id": pid,
            "status": get(r, "status"),
            "industry": get(r, "industry"),
            "title": title,
            "brochure_url": get_brochure_url(ws_values, ws_formulas, r, field_to_col.get("brochure_url")),
            "company_type": get(r, "company_type"),
            "city": get(r, "city"),
            "state": get(r, "state"),
            "employer_origin": get(r, "employer_origin"),
            "requirements": get(r, "requirements"),
            "notes": get(r, "notes"),
            "language_level": get(r, "language_level"),
            "start_date": get(r, "start_date"),
            "salary_pretax": money(get(r, "salary_pretax")),
            "salary_posttax": money(get(r, "salary_posttax")),
            "benefits_notes": get(r, "benefits_notes"),
            "housing": get(r, "housing"),
            "rent_price": get(r, "rent_price"),
            "working_hours": get(r, "working_hours"),
            "day_off": get(r, "day_off"),
            "vacation_days": clean_num(get(r, "vacation_days")),
            "cost_of_living": get(r, "cost_of_living"),
            "category": category_of(get(r, "industry")),
        })
    return rows


def inject_positions(html_path, rows):
    html = html_path.read_text(encoding="utf-8")
    js_array = json.dumps(rows, ensure_ascii=False, indent=4)
    pattern = re.compile(r"const POSITIONS\s*=\s*\[.*?\];", re.S)
    if not pattern.search(html):
        raise ValueError(f"No 'const POSITIONS = [...]' block found in {html_path}")
    new_html = pattern.sub(lambda m: f"const POSITIONS = {js_array};", html, count=1)
    html_path.write_text(new_html, encoding="utf-8")


def main():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--xlsx", required=True, help="Path to the exported Google Sheet (.xlsx)")
    args = parser.parse_args()

    xlsx_path = Path(args.xlsx)
    if not xlsx_path.exists():
        print(f"File not found: {xlsx_path}", file=sys.stderr)
        sys.exit(1)

    wb_values = openpyxl.load_workbook(xlsx_path, data_only=True)
    wb_formulas = openpyxl.load_workbook(xlsx_path, data_only=False)

    # Ausbildung
    if AUSBILDUNG_SHEET in wb_values.sheetnames and AUSBILDUNG_HTML.exists():
        rows = extract_ausbildung_rows(wb_values[AUSBILDUNG_SHEET])
        inject_positions(AUSBILDUNG_HTML, rows)
        print(f"{AUSBILDUNG_SHEET}: {len(rows)} positions -> {AUSBILDUNG_HTML}")
    else:
        print(f"Warning: skipped {AUSBILDUNG_SHEET} (sheet or dashboard file missing)", file=sys.stderr)

    # §18b, 19c, 16d
    if FACHKRAFT_SHEET in wb_values.sheetnames and FACHKRAFT_HTML.exists():
        rows = extract_fachkraft_rows(wb_values[FACHKRAFT_SHEET], wb_formulas[FACHKRAFT_SHEET])
        inject_positions(FACHKRAFT_HTML, rows)
        print(f"{FACHKRAFT_SHEET}: {len(rows)} positions -> {FACHKRAFT_HTML}")
    else:
        print(f"Warning: skipped {FACHKRAFT_SHEET} (sheet or dashboard file missing)", file=sys.stderr)

    print("Done. Redeploy the site folder to publish the changes.")


if __name__ == "__main__":
    main()
